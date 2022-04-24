# reference code from https://python.plainenglish.io/how-to-build-a-whatsapp-chatbot-using-python-and-twilio-ec9367e058f2

from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from twilio.rest import Client
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from datetime import date
from datetime import time
from dateutil.parser import parse
import pytz
import apscheduler
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger


app = Flask(__name__)
count = 0
scheduler = BackgroundScheduler()

# !!!
# need to sync the workbook read/write
# !!!

wb = Workbook()
ws = wb.active
wb.save("remind.xlsx")

def wa_notify(dt, msg):
    account_sid = '<>'
    auth_token = '<>'
    client = Client(account_sid, auth_token)

    message = client.messages.create(
                              from_='whatsapp:+14155238886',
                              body='Reminder for {} for today {}'.format(msg, dt),
                              to='whatsapp:+919900000000'
                          )
    print("sent msg {} on {}".format(msg, dt))


# !!!
# this keeps sending the notifys and DOES NOT check whether it was already sent
# !!!
def wa_sched():
    print("inside wa_notify")

    dt=date.today().strftime('%d/%m/%Y')
    now_date=datetime.strptime(dt,'%d/%m/%Y')
    rem_day=now_date.day
    rem_month=now_date.month
    rem_year=now_date.year

    t=datetime(rem_year,rem_month,rem_day,9,0)
    local = pytz.timezone("Asia/Kolkata")
    local_dt = local.localize(t, is_dst=None)
    utc_dt = local_dt.astimezone(pytz.utc)

    r_wb = load_workbook("remind.xlsx")
    r_ws = r_wb.active

    tot_rows = r_ws.max_row

    if r_ws["A1"] in ['', None]:
        return

    for rn in range(tot_rows):
        c1 = "A{}".format(rn+1)
        c2 = "B{}".format(rn+1)
        dt = r_ws[c1].value
        msg = r_ws[c2].value

        print("{} {} dt is {}, msg is {}".format(c1, c2, dt, msg))

        if dt in [None, '']:
            break

        p= parse(dt)
        parsed_date=p.strftime('%d/%m/%Y')
        
        if parsed_date == dt:
            #scheduler.add_job(send_rem, 'date', run_date=utc_dt, args=[row[0],row[1]])
            wa_notify(dt, msg)
        else:
            pass


def save_reminder_date(date):
    ws = wb.active
    row_filled = ws.max_row
    cell = "A{}".format(row_filled)
    ws[cell] = date
    wb.save("remind.xlsx")
    #sheet.update_cell(row_filled+1, 1, date)
    print("saved date!")
    return 0


def save_reminder_body(msg):
    ws = wb.active
    row_filled = ws.max_row
    cell = "B{}".format(row_filled)
    ws[cell] = msg
    wb.save("remind.xlsx")
    #sheet.update_cell(row_filled+1, 2, msg)
    print("saved reminder message!")
    return 0


def set_reminder_date(msg):
    p= parse(msg)
    date=p.strftime('%d/%m/%Y')
    save_reminder_date(date)
    return 0


def set_reminder_body(msg):
    save_reminder_body(msg)
    return 0


@app.route("/wa", methods=['POST'])
def wa_reply():
    
    incoming_msg = request.form.get('Body').lower()
    response = MessagingResponse()
    message=response.message()
    responded = False
    words = incoming_msg.split('@')

    if "hello" in incoming_msg or "hi" in incoming_msg:
        reply = "Hello! \nDo you want to set a reminder?"
        message.body(reply)
        responded = True

    if len(words) == 1 and "yes" in incoming_msg:
        reminder_string = "Please provide date in the following format only.\n\n"\
        "*Date @* _type the date_ "
        message.body(reminder_string)
        responded = True

    if len(words) == 1 and "no" in incoming_msg:
        reply="Ok. Have a nice day!"
        message.body(reply)
        responded = True
    
    elif len(words) != 1:
        input_type = words[0].strip().lower()
        input_string = words[1].strip()

        if input_type == "date":
            reply="Please enter the reminder message in the following format only.\n\n"\
            "*Reminder @* _type the message_"
            set_reminder_date(input_string)
            message.body(reply)
            responded = True

        if input_type == "reminder":
            reply="Your reminder is set!"
            set_reminder_body(input_string)
            message.body(reply)
            responded = True
        
    if not responded:
        message.body('Incorrect request format. Please enter in the correct format')
    
    return str(response)


@app.route("/sts", methods=["GET", "POST"])
def status_whatsapp():
    #print("STS CALLBACK")
    #print(request.get_data())

    response = MessagingResponse()
    msg = response.message("status")
    return str(response)

  
if __name__ == "__main__":
    #app.run(debug=True)
    scheduler.add_job(func=wa_sched,
                      kwargs={},
                      trigger=IntervalTrigger(seconds=60),
                      id='reminder',
                      name='Initiate notifications periodically',
                      replace_existing=True)

    scheduler.start()

    app.run(debug=False, use_reloader=True, host='0.0.0.0', threaded=True)
